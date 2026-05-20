"""
============================================================
  SYNC STATION — Discord Bot
  Tracks work sessions and generates monthly Excel reports
  Compatible with: Python 3, Zorin OS / Linux
============================================================

COMMANDS:
  !sync in      — Clock in (start your work session)
  !sync break   — Start a break
  !sync back    — Return from break
  !sync out     — Clock out (end your work session)
  !sync status  — See your current session info
  !sync report  — Generate monthly Excel report (admin only)

QUOTA:
  Weekdays = 1 hour active work
  Weekends = 4 hours active work

SCORING:
  Score = (days quota met / total days member synced) * 100
"""

# ============================================================
# IMPORTS
# ============================================================

import discord                          # Discord bot library
from discord.ext import commands        # Command framework
from dotenv import load_dotenv          # Load token from .env file
import os                               # Access environment variables
import json                             # Save/load data as JSON file
from pathlib import Path                # Resolve the repo-local .env path
from datetime import datetime, date     # Handle dates and times
import openpyxl                         # Build Excel (.xlsx) files
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side
)
from openpyxl.utils import get_column_letter
import io                               # Send file in memory (no temp file needed)
import calendar                         # Check weekday/weekend

import threading
from http.server import HTTPServer, BaseHTTPRequestHandler

class PingHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Sync Station is running!")
    def log_message(self, *args):
        pass  # Silence HTTP logs

def run_webserver():
    server = HTTPServer(("0.0.0.0", 8080), PingHandler)
    server.serve_forever()

# Start the web server in a background thread
threading.Thread(target=run_webserver, daemon=True).start()
# ============================================================
# LOAD TOKEN FROM .env FILE
# ============================================================
# Make sure your .env file contains:  TOKEN=your_bot_token_here

load_dotenv(dotenv_path=Path(__file__).resolve().parent / ".env")
TOKEN = os.getenv("TOKEN")

# ============================================================
# BOT SETUP
# ============================================================
# command_prefix = the character before every command (!sync)
# intents = permissions the bot needs

intents = discord.Intents.default()
intents.message_content = True          # Needed to read message text
intents.members = True                  # Needed to list server members

bot = commands.Bot(command_prefix="!", intents=intents)

# ============================================================
# DATA FILE
# ============================================================
# All session data is stored in this JSON file in the same
# folder as bot.py. It is created automatically if missing.

DATA_FILE = "sync_data.json"

# ============================================================
# DATA HELPERS — load and save the JSON file
# ============================================================

def load_data():
    """Load all session data from the JSON file.
    Returns an empty dict if the file doesn't exist yet."""
    if not os.path.exists(DATA_FILE):
        return {}
    with open(DATA_FILE, "r") as f:
        return json.load(f)


def save_data(data):
    """Save all session data back to the JSON file."""
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)


# ============================================================
# DATA STRUCTURE (how each member's data looks in the JSON)
# ============================================================
#
# data = {
#   "123456789" (user ID as string): {
#     "username": "Alice",
#     "sessions": [
#       {
#         "date": "2025-05-01",          # date of session
#         "clock_in": "09:00:00",        # time synced in
#         "clock_out": "17:00:00",       # time synced out (or None)
#         "breaks": [
#           {"start": "12:00:00", "end": "12:30:00"},
#           ...
#         ],
#         "active_seconds": 27000,       # total work time minus breaks
#         "quota_met": true              # did they hit the daily quota?
#       },
#       ...
#     ]
#   }
# }

# ============================================================
# TIME HELPERS
# ============================================================

def now_str():
    """Return current time as HH:MM:SS string."""
    return datetime.now().strftime("%H:%M:%S")


def today_str():
    """Return today's date as YYYY-MM-DD string."""
    return date.today().isoformat()


def seconds_to_hms(seconds):
    """Convert a number of seconds into a readable H h M m string."""
    seconds = int(seconds)
    h = seconds // 3600
    m = (seconds % 3600) // 60
    return f"{h}h {m}m"


def daily_quota_seconds(date_str):
    """Return the required active seconds for a given date.
    Weekday = 1 hour (3600s), Weekend = 4 hours (14400s)."""
    d = date.fromisoformat(date_str)
    if d.weekday() >= 5:        # Saturday = 5, Sunday = 6
        return 4 * 3600         # 4 hours for weekends
    return 1 * 3600             # 1 hour for weekdays


def calc_active_seconds(session):
    """Calculate total active (non-break) seconds for a session.
    Works even if the user hasn't clocked out yet (uses current time)."""
    fmt = "%H:%M:%S"

    clock_in = datetime.strptime(session["clock_in"], fmt)
    clock_out_str = session.get("clock_out")
    clock_out = (
        datetime.strptime(clock_out_str, fmt)
        if clock_out_str
        else datetime.now().replace(microsecond=0)
    )

    total = (clock_out - clock_in).total_seconds()

    # Subtract all completed break durations
    for b in session.get("breaks", []):
        if b.get("end"):
            b_start = datetime.strptime(b["start"], fmt)
            b_end = datetime.strptime(b["end"], fmt)
            total -= (b_end - b_start).total_seconds()

    return max(0, total)        # Never return negative seconds


# ============================================================
# GET OR CREATE MEMBER RECORD
# ============================================================

def get_member(data, user):
    """Return (and create if needed) the data record for a user."""
    uid = str(user.id)
    if uid not in data:
        data[uid] = {
            "username": user.display_name,
            "sessions": []
        }
    else:
        # Always keep the username up to date
        data[uid]["username"] = user.display_name
    return data[uid]


def get_today_session(member_data):
    """Return today's session dict for a member, or None if not started."""
    today = today_str()
    for s in member_data["sessions"]:
        if s["date"] == today:
            return s
    return None


# ============================================================
# BOT EVENTS
# ============================================================

@bot.event
async def on_ready():
    """Runs when the bot successfully connects to Discord."""
    print(f"✅ Sync Station is online as {bot.user}")
    print(f"   Tracking data in: {os.path.abspath(DATA_FILE)}")


# ============================================================
# MAIN COMMAND GROUP: !sync
# ============================================================

@bot.group(name="sync", invoke_without_command=True)
async def sync_group(ctx):
    """Shows help if !sync is used without a sub-command."""
    embed = discord.Embed(
        title="🕐 Sync Station",
        description="Track your daily work sessions.",
        color=discord.Color.blurple()
    )
    embed.add_field(name="!sync in",     value="Clock in",               inline=False)
    embed.add_field(name="!sync break",  value="Start a break",          inline=False)
    embed.add_field(name="!sync back",   value="Return from break",      inline=False)
    embed.add_field(name="!sync out",    value="Clock out",              inline=False)
    embed.add_field(name="!sync status", value="See your session info",  inline=False)
    embed.add_field(name="!sync report", value="Monthly Excel report (admin)", inline=False)
    await ctx.send(embed=embed)


# ============================================================
# !sync in — Clock in
# ============================================================

@sync_group.command(name="in")
async def sync_in(ctx):
    """Clock in for the day. Creates a new session for today."""
    data = load_data()
    member = get_member(data, ctx.author)
    today = today_str()

    # Prevent double clock-in on the same day
    if get_today_session(member):
        await ctx.send(f"⚠️ {ctx.author.mention} you already synced in today!")
        return

    # Create a new session entry for today
    session = {
        "date": today,
        "clock_in": now_str(),
        "clock_out": None,
        "breaks": [],
        "active_seconds": 0,
        "quota_met": False
    }
    member["sessions"].append(session)
    save_data(data)

    quota = daily_quota_seconds(today)
    quota_label = "1 hour" if quota == 3600 else "4 hours"

    await ctx.send(
        f"✅ **{ctx.author.display_name}** synced in at `{session['clock_in']}`\n"
        f"📋 Today's quota: **{quota_label}** of active time"
    )


# ============================================================
# !sync break — Start a break
# ============================================================

@sync_group.command(name="break")
async def sync_break(ctx):
    """Start a break. Must be clocked in and not already on break."""
    data = load_data()
    member = get_member(data, ctx.author)
    session = get_today_session(member)

    if not session:
        await ctx.send(f"⚠️ {ctx.author.mention} you haven't synced in today. Use `!sync in` first.")
        return

    if session.get("clock_out"):
        await ctx.send(f"⚠️ {ctx.author.mention} you already synced out for today.")
        return

    # Check if already on an open break (a break with no end time)
    for b in session["breaks"]:
        if not b.get("end"):
            await ctx.send(f"⚠️ {ctx.author.mention} you're already on a break!")
            return

    # Open a new break
    session["breaks"].append({"start": now_str(), "end": None})
    save_data(data)

    await ctx.send(f"☕ **{ctx.author.display_name}** started a break at `{session['breaks'][-1]['start']}`")


# ============================================================
# !sync back — Return from break
# ============================================================

@sync_group.command(name="back")
async def sync_back(ctx):
    """Return from a break. Closes the open break timer."""
    data = load_data()
    member = get_member(data, ctx.author)
    session = get_today_session(member)

    if not session:
        await ctx.send(f"⚠️ {ctx.author.mention} you haven't synced in today.")
        return

    # Find the open break (one without an end time)
    open_break = None
    for b in session["breaks"]:
        if not b.get("end"):
            open_break = b
            break

    if not open_break:
        await ctx.send(f"⚠️ {ctx.author.mention} you're not on a break right now.")
        return

    # Close the break
    open_break["end"] = now_str()
    save_data(data)

    fmt = "%H:%M:%S"
    break_secs = (
        datetime.strptime(open_break["end"], fmt) -
        datetime.strptime(open_break["start"], fmt)
    ).total_seconds()

    await ctx.send(
        f"💪 **{ctx.author.display_name}** is back from break at `{open_break['end']}`\n"
        f"   Break duration: {seconds_to_hms(break_secs)}"
    )


# ============================================================
# !sync out — Clock out
# ============================================================

@sync_group.command(name="out")
async def sync_out(ctx):
    """Clock out for the day. Calculates active time and checks quota."""
    data = load_data()
    member = get_member(data, ctx.author)
    session = get_today_session(member)

    if not session:
        await ctx.send(f"⚠️ {ctx.author.mention} you haven't synced in today.")
        return

    if session.get("clock_out"):
        await ctx.send(f"⚠️ {ctx.author.mention} you already synced out today.")
        return

    # Auto-close any open break before clocking out
    for b in session["breaks"]:
        if not b.get("end"):
            b["end"] = now_str()

    # Record clock-out time and calculate final active seconds
    session["clock_out"] = now_str()
    active = calc_active_seconds(session)
    session["active_seconds"] = active

    # Check if the daily quota was met
    quota = daily_quota_seconds(session["date"])
    session["quota_met"] = active >= quota

    save_data(data)

    quota_label = "1 hour" if quota == 3600 else "4 hours"
    status = "✅ Quota met!" if session["quota_met"] else f"❌ Quota not met (needed {quota_label})"

    await ctx.send(
        f"👋 **{ctx.author.display_name}** synced out at `{session['clock_out']}`\n"
        f"⏱️ Active time today: **{seconds_to_hms(active)}**\n"
        f"{status}"
    )


# ============================================================
# !sync status — Show today's session info
# ============================================================

@sync_group.command(name="status")
async def sync_status(ctx):
    """Show the current session status for the user."""
    data = load_data()
    member = get_member(data, ctx.author)
    session = get_today_session(member)

    if not session:
        await ctx.send(f"📋 {ctx.author.mention} you haven't synced in today. Use `!sync in` to start.")
        return

    active = calc_active_seconds(session)
    quota = daily_quota_seconds(session["date"])
    remaining = max(0, quota - active)

    # Detect if currently on a break
    on_break = any(not b.get("end") for b in session["breaks"])

    clocked_out = "Yes" if session.get("clock_out") else ("On break ☕" if on_break else "No")

    embed = discord.Embed(
        title=f"📊 {ctx.author.display_name}'s Session — {session['date']}",
        color=discord.Color.green() if active >= quota else discord.Color.orange()
    )
    embed.add_field(name="Clocked in",   value=session["clock_in"],         inline=True)
    embed.add_field(name="Clocked out",  value=session.get("clock_out") or clocked_out, inline=True)
    embed.add_field(name="Active time",  value=seconds_to_hms(active),      inline=True)
    embed.add_field(name="Quota",        value=seconds_to_hms(quota),       inline=True)
    embed.add_field(name="Remaining",    value=seconds_to_hms(remaining),   inline=True)
    embed.add_field(name="Quota met",    value="✅ Yes" if active >= quota else "❌ Not yet", inline=True)
    embed.add_field(name="Breaks taken", value=str(len(session["breaks"])),  inline=True)

    await ctx.send(embed=embed)


# ============================================================
# !sync report — Generate monthly Excel report
# ============================================================

@sync_group.command(name="report")
@commands.has_permissions(administrator=True)   # Only admins can run this
async def sync_report(ctx, month: int = None, year: int = None):
    """Generate and send a monthly Excel report.
    Usage: !sync report          (current month)
           !sync report 4 2025   (April 2025)
    """
    # Default to current month/year if not specified
    today = date.today()
    if month is None:
        month = today.month
    if year is None:
        year = today.year

    month_name = calendar.month_name[month]
    await ctx.send(f"⏳ Generating report for **{month_name} {year}**...")

    data = load_data()

    if not data:
        await ctx.send("❌ No sync data found yet.")
        return

    # --------------------------------------------------------
    # Figure out how many days were in this month and
    # how many were weekdays vs weekends
    # --------------------------------------------------------
    days_in_month = calendar.monthrange(year, month)[1]
    month_dates = [date(year, month, d) for d in range(1, days_in_month + 1)]

    # Only count days up to today (don't penalise future days)
    cutoff = min(today, date(year, month, days_in_month))
    past_dates = [d for d in month_dates if d <= cutoff]

    total_days = len(past_dates)        # Total days counted so far

    # --------------------------------------------------------
    # Build per-member stats
    # --------------------------------------------------------
    rows = []

    for uid, member in data.items():
        username = member.get("username", f"User {uid}")

        # Collect this member's sessions that fall in the chosen month
        month_sessions = [
            s for s in member["sessions"]
            if s["date"].startswith(f"{year}-{month:02d}")
        ]

        days_synced   = len(month_sessions)
        days_quota_met = sum(1 for s in month_sessions if s.get("quota_met"))
        total_active   = sum(s.get("active_seconds", 0) for s in month_sessions)

        # Score = quota days met / total past days in month * 100
        # A member who wasn't online on a day gets 0 for that day
        score = round((days_quota_met / total_days) * 100, 1) if total_days > 0 else 0

        rows.append({
            "username":       username,
            "days_synced":    days_synced,
            "days_quota_met": days_quota_met,
            "total_days":     total_days,
            "total_active":   seconds_to_hms(total_active),
            "score":          score
        })

    # Sort by score descending (highest first)
    rows.sort(key=lambda r: r["score"], reverse=True)

    # --------------------------------------------------------
    # Build the Excel file with openpyxl
    # --------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name} {year}"

    # --- Colour palette ---
    COLOR_HEADER_BG  = "1E3A5F"     # Dark navy
    COLOR_HEADER_FG  = "FFFFFF"     # White
    COLOR_TITLE_BG   = "2E86AB"     # Teal blue
    COLOR_ALT_ROW    = "EEF4FB"     # Light blue-grey
    COLOR_GOLD       = "F6C90E"     # Gold for top scorer
    COLOR_GREEN_BG   = "D6F5D6"     # Light green for 80%+
    COLOR_ORANGE_BG  = "FFF3CD"     # Light orange for 50-79%
    COLOR_RED_BG     = "FDECEA"     # Light red for <50%

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Title row ---
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Sync Station — Monthly Report: {month_name} {year}"
    title_cell.font      = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
    title_cell.fill      = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Subtitle info row ---
    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = (
        f"Total days counted: {total_days}  |  "
        f"Quota: 1hr weekdays / 4hr weekends  |  "
        f"Generated: {today.strftime('%d %b %Y')}"
    )
    sub_cell.font      = Font(name="Arial", size=10, color="555555")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # --- Column headers (row 3) ---
    headers = [
        "Rank", "Member", "Days Synced",
        "Days Quota Met", "Total Days", "Active Time", "Score / 100"
    ]
    col_widths = [6, 22, 14, 16, 12, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = Font(name="Arial", bold=True, size=11, color=COLOR_HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22

    # --- Data rows ---
    for rank, row in enumerate(rows, start=1):
        excel_row = rank + 3       # Rows 4 onwards

        score = row["score"]

        # Pick row background colour based on score
        if rank == 1:
            row_bg = COLOR_GOLD         # Top scorer gets gold
        elif score >= 80:
            row_bg = COLOR_GREEN_BG
        elif score >= 50:
            row_bg = COLOR_ORANGE_BG
        else:
            row_bg = COLOR_RED_BG

        # Alternate light blue for even rows (when not coloured by score)
        if rank % 2 == 0 and rank != 1 and score >= 80:
            row_bg = COLOR_ALT_ROW

        values = [
            rank,
            row["username"],
            row["days_synced"],
            row["days_quota_met"],
            row["total_days"],
            row["total_active"],
            score
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=11,
                                  bold=(col_idx == 7),    # Bold the score column
                                  color="1E3A5F" if rank == 1 else "000000")
            cell.fill      = PatternFill("solid", fgColor=row_bg)
            cell.alignment = Alignment(
                horizontal="center" if col_idx != 2 else "left",
                vertical="center"
            )
            cell.border = border

        ws.row_dimensions[excel_row].height = 20

    # --- Freeze top 3 rows so header scrolls with data ---
    ws.freeze_panes = "A4"

    # --------------------------------------------------------
    # Send the Excel file directly into Discord
    # --------------------------------------------------------
    # We use io.BytesIO so no temp file is written to disk
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    filename = f"sync_report_{month_name}_{year}.xlsx"
    discord_file = discord.File(fp=excel_buffer, filename=filename)

    # Build a summary embed to accompany the file
    embed = discord.Embed(
        title=f"📊 Sync Station Report — {month_name} {year}",
        color=discord.Color.blurple()
    )
    embed.add_field(name="Members tracked", value=str(len(rows)),  inline=True)
    embed.add_field(name="Days counted",    value=str(total_days), inline=True)

    if rows:
        top = rows[0]
        embed.add_field(
            name="🥇 Top scorer",
            value=f"{top['username']} — {top['score']}/100",
            inline=False
        )

    # List every member's score in the embed
    score_lines = "\n".join(
        f"{'🥇' if i == 0 else '▸'} **{r['username']}** — {r['score']}/100"
        for i, r in enumerate(rows)
    )
    if score_lines:
        embed.add_field(name="All scores", value=score_lines, inline=False)

    embed.set_footer(text=f"Generated by Sync Station • {today.strftime('%d %b %Y')}")

    await ctx.send(embed=embed, file=discord_file)


# ============================================================
# ERROR HANDLERS
# ============================================================

@sync_report.error
async def report_error(ctx, error):
    """Handle errors for the !sync report command."""
    if isinstance(error, commands.MissingPermissions):
        await ctx.send("❌ Only server administrators can generate reports.")
    else:
        await ctx.send(f"❌ An error occurred: {error}")


@bot.event
async def on_command_error(ctx, error):
    """Handle general command errors."""
    if isinstance(error, commands.CommandNotFound):
        pass    # Silently ignore unknown commands
    elif isinstance(error, commands.MissingRequiredArgument):
        await ctx.send(f"⚠️ Missing argument. Try `!sync` for help.")
    else:
        print(f"Error: {error}")


# ============================================================
# RUN THE BOT
# ============================================================
# This is the last line — it starts the bot using your token.

bot.run(TOKEN)